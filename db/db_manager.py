import os
import pymysql
import pymysql.cursors
from datetime import datetime
from werkzeug.utils import secure_filename
# for Excel download
import pandas as pd
import openpyxl
from flask import current_app, send_file
from io import BytesIO
from services.analyze_environmental import analyze_environmental_aspects


# ===================== 공통 DB 연결 =====================
def get_connection():
    return pymysql.connect(
        host="localhost",
        user="root",
        password="0000",  # ← 수정: DB 비밀번호 변경
        database="garbageguard",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor
    )

# ===================== Company 관련 =====================
# ✅ 회사 등록
def insert_company(company_name, address, ceo_name, contact):
    conn = get_connection()
    with conn.cursor() as cursor:
        sql = """
            INSERT INTO companies (company_name, address, ceo_name, contact)
            VALUES (%s, %s, %s, %s)
        """
        cursor.execute(sql, (company_name, address, ceo_name, contact))
        conn.commit()
    conn.close()

# ✅ 회사 삭제
def delete_company(company_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "DELETE FROM companies WHERE company_id = %s",
            (company_id,)
        )
        conn.commit()
    conn.close()

# ✅ 회사 목록 조회
def get_all_companies():
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("SELECT company_id, company_name FROM companies ORDER BY company_name")
        companies = cursor.fetchall()
    conn.close()
    return companies

# ✅ 회사 ID로 회사 정보 조회
def get_company_by_id(company_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT company_id, company_name FROM companies WHERE company_id=%s",
            (company_id,)
        )
        result = cursor.fetchone()
    conn.close()
    return result

# ✅ 회사명으로 company_id 조회 및 해당 회사의 건설현장 리스트 조회
def get_sites_by_company(company_name):
    conn = get_connection()
    with conn.cursor() as cursor:
        sql = """
            SELECT cs.site_name
              FROM construction_sites cs
              JOIN companies c ON cs.company_id = c.company_id
             WHERE c.company_name = %s
        """
        cursor.execute(sql, (company_name,))
        rows = cursor.fetchall()
    conn.close()
    return [row['site_name'] for row in rows]

def get_all_departments():
    """
    부서 목록을 조회해서 [{'id': ..., 'name': ...}, …] 형태로 반환합니다.
    사전 테이블 departments(dept_id, dept_name)가 필요합니다.
    """
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("""
            SELECT dept_id AS id,
                   dept_name AS name
              FROM departments
             ORDER BY name
        """)
        rows = cursor.fetchall()
    conn.close()
    return rows


# ===================== Construction Site 관련 =====================
# ✅ 건설 현장 등록
def upload_construction_site(
    site_name, address, manager_name,
    latitude=None, longitude=None, company_id=None,
    department=None, importance_level=None,
    contractor_notes=None, calibration_date=None,
    survey_file=None, procedure_file=None,
    standard_file=None, monitoring_data=None,
    calibration_file=None
):
    conn = get_connection()
    with conn.cursor() as cursor:
        # 기본 필드 삽입 준비
        columns = ["site_name", "address", "manager_name"]
        values  = [site_name, address, manager_name]

        # 추가 필드
        columns += [
            "company_id", "department",
            "importance_level", "contractor_notes",
            "calibration_date"
        ]
        values += [
            company_id, department,
            importance_level, contractor_notes,
            calibration_date
        ]

        # 좌표 있을 경우
        if latitude is not None and longitude is not None:
            columns += ["latitude", "longitude"]
            values += [latitude, longitude]

        # INSERT 실행
        placeholders = ", ".join(["%s"] * len(columns))
        sql = f"INSERT INTO construction_sites ({', '.join(columns)}) VALUES ({placeholders})"
        cursor.execute(sql, tuple(values))
        new_id = cursor.lastrowid

        # ✅ UPLOAD_DIR 설정 확인
        upload_dir = current_app.config.get("UPLOAD_DIR")
        if not upload_dir:
            raise ValueError("UPLOAD_DIR 설정이 없습니다.")

        # ✅ 파일 업로드 처리
        for field_name, file_obj in [
            ("survey_file_path", survey_file),
            ("procedure_file_path", procedure_file),
            ("standard_file_path", standard_file),
            ("monitoring_data_path", monitoring_data),
            ("calibration_file_path", calibration_file)
        ]:
            if file_obj and file_obj.filename:
                filename = secure_filename(file_obj.filename)
                save_path = os.path.join(upload_dir, filename)
                file_obj.save(save_path)

                cursor.execute(
                    f"UPDATE construction_sites SET {field_name}=%s WHERE site_id=%s",
                    (save_path, new_id)
                )

        conn.commit()
    conn.close()

# db/db_manager.py

def get_site_info(company_name: str, site_name: str) -> dict:
    """
    부서·중요도·메모·교정일자 뿐 아니라,
    survey_file_path, procedure_file_path, standard_file_path,
    monitoring_data_path, calibration_file_path 까지 반환합니다.
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT
                    cs.department,
                    cs.importance_level,
                    cs.contractor_notes,
                    DATE_FORMAT(cs.calibration_date, '%%Y-%%m-%%d') AS calibration_date,
                    cs.survey_file_path,
                    cs.procedure_file_path,
                    cs.standard_file_path,
                    cs.monitoring_data_path,
                    cs.calibration_file_path
                FROM construction_sites cs
                JOIN companies c ON cs.company_id = c.company_id
                WHERE c.company_name = %s
                  AND cs.site_name    = %s
                LIMIT 1
            """, (company_name, site_name))
            row = cursor.fetchone()
    finally:
        conn.close()

    if not row:
        return {
            'department': None,
            'importance_level': None,
            'contractor_notes': None,
            'calibration_date': None,
            'survey_file_path': None,
            'procedure_file_path': None,
            'standard_file_path': None,
            'monitoring_data_path': None,
            'calibration_file_path': None
        }

    return {
        'department':           row.get('department'),
        'importance_level':     row.get('importance_level'),
        'contractor_notes':     row.get('contractor_notes'),
        'calibration_date':     row.get('calibration_date'),
        'survey_file_path':     row.get('survey_file_path'),
        'procedure_file_path':  row.get('procedure_file_path'),
        'standard_file_path':   row.get('standard_file_path'),
        'monitoring_data_path': row.get('monitoring_data_path'),
        'calibration_file_path':row.get('calibration_file_path')
    }


# ✅ 건설 현장 수정
def update_construction_site(
    site_id, site_name, address, manager_name,
    latitude=None, longitude=None, company_id=None,
    department=None, importance_level=None,
    contractor_notes=None, calibration_date=None,
    survey_file=None, procedure_file=None,
    standard_file=None, monitoring_data=None,
    calibration_file=None
):
    conn = get_connection()
    with conn.cursor() as cursor:
        # 기본 필드 업데이트 준비
        updates = ["site_name=%s", "address=%s", "manager_name=%s"]
        values  = [site_name, address, manager_name]

        # 추가 필드들
        updates += [
            "company_id=%s", "department=%s",
            "importance_level=%s", "contractor_notes=%s",
            "calibration_date=%s"
        ]
        values += [
            company_id, department,
            importance_level, contractor_notes,
            calibration_date
        ]

        # 좌표 있을 경우 추가
        if latitude is not None and longitude is not None:
            updates += ["latitude=%s", "longitude=%s"]
            values += [latitude, longitude]

        # 최종 UPDATE 쿼리 실행
        sql = f"UPDATE construction_sites SET {', '.join(updates)} WHERE site_id=%s"
        values.append(site_id)
        cursor.execute(sql, tuple(values))

        # ✅ 파일 업로드 디렉토리 설정 확인
        upload_dir = current_app.config.get("UPLOAD_DIR")
        if not upload_dir:
            raise ValueError("UPLOAD_DIR 설정이 없습니다.")

        # ✅ 파일 업로드 및 경로 업데이트
        for field_name, file_obj in [
            ("survey_file_path", survey_file),
            ("procedure_file_path", procedure_file),
            ("standard_file_path", standard_file),
            ("monitoring_data_path", monitoring_data),
            ("calibration_file_path", calibration_file)
        ]:
            if file_obj and file_obj.filename:
                filename = secure_filename(file_obj.filename)
                save_path = os.path.join(upload_dir, filename)
                file_obj.save(save_path)

                cursor.execute(
                    f"UPDATE construction_sites SET {field_name}=%s WHERE site_id=%s",
                    (save_path, site_id)
                )

        conn.commit()
    conn.close()

# ✅ 건설 현장 삭제
def delete_construction_site(site_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("DELETE FROM construction_sites WHERE site_id=%s", (site_id,))
        conn.commit()
    conn.close()

# ✅ 현장 전체 목록 조회
def get_all_construction_sites():
    """
    construction_sites 테이블에서
    - 기본 텍스트/숫자 컬럼
    - DATE_FORMAT(calibration_date, '%Y-%m-%d') AS calibration_date
    - 파일 경로 컬럼 5종 (survey|procedure|standard|monitoring|calibration)_file_path
    를 모두 SELECT 해서 반환합니다.
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT
                    site_id,
                    site_name,
                    address,
                    manager_name,
                    latitude,
                    longitude,
                    company_id,
                    department,
                    importance_level,
                    contractor_notes,
                    DATE_FORMAT(calibration_date, '%Y-%m-%d') AS calibration_date,
                    survey_file_path,
                    procedure_file_path,
                    standard_file_path,
                    monitoring_data_path,
                    calibration_file_path
                FROM construction_sites
            """)
            result = cursor.fetchall()
    finally:
        conn.close()
    return result

# alias for site list
def get_all_sites():
    return get_all_construction_sites()

# ✅ 특정 현장 ID로 현장 정보 조회
def get_site_by_id(site_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT site_id, site_name, company_id FROM construction_sites WHERE site_id=%s",
            (site_id,)
        )
        result = cursor.fetchone()
    conn.close()
    return result

# ✅ site_name → site_id 변환
def get_site_id_by_name(site_name):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT site_id FROM construction_sites WHERE site_name=%s",
            (site_name,)
        )
        row = cursor.fetchone()
    conn.close()
    return row['site_id'] if row else None

# ===================== Waste Photo & Management 관련 =====================
# ✅ YOLO 결과 저장
def insert_waste_photo(site_id, object_id, image_filename, detection_summary, uploaded_at=None):
    conn = get_connection()
    with conn.cursor() as cursor:
        if uploaded_at is None:
            sql = """
                INSERT INTO waste_photos (site_id, object_id, image_filename, detection_summary, is_analyzed)
                VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (site_id, object_id, image_filename, detection_summary, True))
        else:
            sql = """
                INSERT INTO waste_photos (site_id, object_id, image_filename, detection_summary, uploaded_at, is_analyzed)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (site_id, object_id, image_filename, detection_summary, uploaded_at, True))
        conn.commit()
    conn.close()

# ✅ detection_summary 조회
def get_detection_summary(site_id, image_filename):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT detection_summary FROM waste_photos WHERE site_id=%s AND image_filename=%s",
            (site_id, image_filename)
        )
        result = cursor.fetchone()
    conn.close()
    return result['detection_summary'] if result else None

# ✅ 특정 날짜의 현장 이미지 조회
def get_photos_by_site_and_date(site_id, date):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT image_filename, detection_summary, uploaded_at FROM waste_photos WHERE site_id=%s AND DATE(uploaded_at)=%s ORDER BY uploaded_at DESC",
            (site_id, date)
        )
        result = cursor.fetchall()
    conn.close()
    return result

# ✅ 특정 현장의 모든 이미지 조회
def get_photos_by_site(site_id):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT image_filename, detection_summary, uploaded_at FROM waste_photos WHERE site_id=%s ORDER BY uploaded_at DESC",
            (site_id,)
        )
        result = cursor.fetchall()
    conn.close()
    return result

# ===================== Object 관련 =====================
# ✅ object_name → object_id 자동 생성 및 조회
def get_object_id_by_name(object_name):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT object_id FROM waste_objects WHERE object_name=%s",
            (object_name,)
        )
        row = cursor.fetchone()
        if row:
            obj_id = row['object_id']
        else:
            cursor.execute(
                "INSERT INTO waste_objects (object_name) VALUES (%s)",
                (object_name,)
            )
            conn.commit()
            obj_id = cursor.lastrowid
    conn.close()
    return obj_id

# ===================== 통계 관련 =====================
# ✅ 월별 통계 조회
def get_monthly_stats(site_id=None):
    conn = get_connection()
    with conn.cursor() as cursor:
        if site_id:
            cursor.execute(
                """
                    SELECT DATE_FORMAT(disposal_date,'%%Y-%%m') AS month,
                           SUM(waste_amount)     AS total_waste,
                           SUM(carbon_emission)  AS total_emission
                      FROM waste_management
                     WHERE site_id=%s
                     GROUP BY month
                     ORDER BY month
                """,
                (site_id,)
            )
        else:
            cursor.execute(
                """
                    SELECT DATE_FORMAT(disposal_date,'%%Y-%%m') AS month,
                           SUM(waste_amount)     AS total_waste,
                           SUM(carbon_emission)  AS total_emission
                      FROM waste_management
                     GROUP BY month
                     ORDER BY month
                """
            )
        rows = cursor.fetchall()
    conn.close()
    return rows


# ✅ 폐기물 관리 저장
def insert_waste_management(site_id, waste_type, waste_amount, carbon_emission, disposal_date):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "INSERT INTO waste_management (site_id, waste_type, waste_amount, carbon_emission, disposal_date) VALUES (%s, %s, %s, %s, %s)",
            (site_id, waste_type, waste_amount, carbon_emission, disposal_date)
        )
        conn.commit()
    conn.close()

# ✅ 한 장의 사진(photo_id)에 대해 탐지된 객체(object_id)와 개수(count) 저장
def insert_photo_object(photo_id: int, object_id: int, count: int):
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
                INSERT INTO waste_photo_objects (photo_id, object_id, count)
                VALUES (%s, %s, %s)
            """
            cursor.execute(sql, (photo_id, object_id, count))
        conn.commit()
    finally:
        conn.close()

# ✅ 폐기물 처리 내역 저장 (통계용, 상세 정보 포함)
def insert_waste_management(
    site_id,
    waste_type,
    waste_amount,
    carbon_emission,
    disposal_date,
    waste_category,
    waste_code,
    recyclable
):
    conn = get_connection()
    with conn.cursor() as cursor:
        sql = """
            INSERT INTO waste_management
              (site_id, waste_type, waste_amount, carbon_emission,
               disposal_date, waste_category, waste_code, recyclable)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (
            site_id,
            waste_type,
            waste_amount,
            carbon_emission,
            disposal_date,
            waste_category,
            waste_code,
            int(recyclable)
        ))
        conn.commit()
    conn.close()

# ▶ 추가 함수 1: 폐기물 종류별 탄소 배출량
def get_emissions_by_waste_type(site_id=None):
    conn = get_connection()
    with conn.cursor() as cursor:
        if site_id:
            sql = """
                SELECT waste_type, SUM(carbon_emission) AS total_emission
                FROM waste_management
                WHERE site_id = %s
                GROUP BY waste_type
            """
            cursor.execute(sql, (site_id,))
        else:
            cursor.execute("""
                SELECT waste_type, SUM(carbon_emission) AS total_emission
                FROM waste_management
                GROUP BY waste_type
            """)
        result = cursor.fetchall()
    conn.close()
    return result

# ▶ 추가 함수 2: 폐기물 종류별 배출량
def get_waste_amount_by_type(site_id=None):
    conn = get_connection()
    with conn.cursor() as cursor:
        if site_id:
            sql = """
                SELECT waste_type, SUM(waste_amount) AS total_amount
                FROM waste_management
                WHERE site_id = %s
                GROUP BY waste_type
            """
            cursor.execute(sql, (site_id,))
        else:
            cursor.execute("""
                SELECT waste_type, SUM(waste_amount) AS total_amount
                FROM waste_management
                GROUP BY waste_type
            """)
        result = cursor.fetchall()
    conn.close()
    return result

# ▶ 추가 함수 3: 현재 월 폐기물 비율 (수정판)
def get_waste_percentage_current_month():
    conn = get_connection()
    with conn.cursor() as cursor:
        now = datetime.now()
        current_month = now.strftime('%Y-%m')
        sql = """
            SELECT waste_type, SUM(waste_amount) AS total
            FROM waste_management
            WHERE DATE_FORMAT(disposal_date, '%%Y-%%m') = %s
            GROUP BY waste_type
        """
        cursor.execute(sql, (current_month,))
        rows = cursor.fetchall()
        total = sum(row['total'] for row in rows)
        if total == 0:
            return []
        return [
            {"waste_type": row["waste_type"], "percentage": round((row["total"] / total) * 100)}
            for row in rows
        ]
    conn.close()

# ▶ 추가 함수 4: 건설사별 탄소 배출량
def get_carbon_emission_by_company():
    conn = get_connection()
    with conn.cursor() as cursor:
        sql = """
            SELECT c.company_name, SUM(wm.carbon_emission) AS total_emission
            FROM companies c
            JOIN construction_sites cs ON c.company_id = cs.company_id
            JOIN waste_management wm ON cs.site_id = wm.site_id
            GROUP BY c.company_id, c.company_name
            ORDER BY total_emission DESC
        """
        cursor.execute(sql)
        result = cursor.fetchall()
    conn.close()
    return result

# ▶ 추가 함수 5: 최다 배출 건설사 조회
def get_top_carbon_emitter_company():
    conn = get_connection()
    with conn.cursor() as cursor:
        sql = """
            SELECT c.company_name, SUM(wm.carbon_emission) AS total_emission
            FROM companies c
            JOIN construction_sites cs ON c.company_id = cs.company_id
            JOIN waste_management wm ON cs.site_id = wm.site_id
            GROUP BY c.company_id, c.company_name
            ORDER BY total_emission DESC
            LIMIT 1
        """
        cursor.execute(sql)
        result = cursor.fetchone()
    conn.close()
    return result

# ▶ 추가 함수 6: 지역별 월별 탄소 배출량
def get_monthly_emission_by_region(region_keyword: str):
    region_keyword = region_keyword.replace("광역시", "").replace("특별시", "").replace("도", "").strip()
    conn = get_connection()
    with conn.cursor(pymysql.cursors.DictCursor) as cursor:  # ← 수정: DictCursor 적용
        sql = (
            "SELECT cs.site_name, "
            "DATE_FORMAT(wm.disposal_date, '%%Y-%%m') AS month, "
            "SUM(COALESCE(wm.carbon_emission,0)) AS total_emission "
            "FROM construction_sites cs "
            "JOIN waste_management wm ON cs.site_id = wm.site_id "
            "WHERE cs.address LIKE %s "
            "GROUP BY cs.site_name, month "
            "ORDER BY cs.site_name, month"
        )
        cursor.execute(sql, (f"%{region_keyword}%",))
        result = cursor.fetchall()
        print(result)  # ← 디버깅용 출력
        for row in result:
            row['total_emission'] = float(row['total_emission'])
    conn.close()
    return result

# ▶ 추가 함수 7: 지역별 상위 폐기물 종류
def get_top_waste_types_by_region(region_keyword: str):
    region_keyword = region_keyword.replace("광역시", "").replace("특별시", "").replace("도", "").strip()
    conn = get_connection()
    with conn.cursor() as cursor:
        sql = """
            SELECT wm.waste_type, SUM(wm.waste_amount) AS total_amount
            FROM construction_sites cs
            JOIN waste_management wm ON cs.site_id = wm.site_id
            WHERE cs.address LIKE %s
            GROUP BY wm.waste_type
            ORDER BY total_amount DESC
            LIMIT 10
        """
        cursor.execute(sql, (f"%{region_keyword}%",))
        result = cursor.fetchall()
        for row in result:
            row['total_amount'] = float(row['total_amount'])
    conn.close()
    return result

# ▶ 추가 함수 8: 월별 폐기물과 탄소 배출량
def get_monthly_stats2(site_id=None):
    conn = get_connection()
    with conn.cursor() as cursor:
        if site_id:
            cursor.execute(
                """
                    SELECT DATE_FORMAT(disposal_date,'%Y년 %m월') AS month,
                           SUM(waste_amount)     AS total_waste,
                           SUM(carbon_emission)  AS total_emission
                      FROM waste_management
                     WHERE site_id=%s
                     GROUP BY month
                     ORDER BY month
                """,
                (site_id,)
            )
        else:
            cursor.execute(
                """
                    SELECT DATE_FORMAT(disposal_date,'%%Y-%%m') AS month,
                           SUM(waste_amount)     AS total_waste,
                           SUM(carbon_emission)  AS total_emission
                      FROM waste_management
                     GROUP BY month
                     ORDER BY month
                """
            )
        rows = cursor.fetchall()
    conn.close()
    return rows



def get_images_for_site(company_name: str, site_name: str) -> list:
    site_id = get_site_id_by_name(site_name)
    if site_id is None:
        return []
    return get_photos_by_site(site_id)

def analyze_waste_requirements(images: list) -> list:
    # 현재 사용 안 함
    return []


def insert_audit_session(company_id: int, site_id: int, performed_by: str) -> int:
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            "INSERT INTO audit_sessions (site_id, company_id, performed_by) VALUES (%s, %s, %s)",
            (site_id, company_id, performed_by)
        )
        session_id = cursor.lastrowid
        conn.commit()
    conn.close()
    return session_id


def insert_audit_result(session_id: int, item_type: str, item_index: int, is_pass: bool, comment: str = ''):
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO audit_results
              (session_id, item_type, item_index, is_pass, comment)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (session_id, item_type, item_index, is_pass, comment)
        )
        conn.commit()
    conn.close()

def create_audit_report_excel(company_name: str, site_name: str) -> BytesIO:
    tpl_path = os.path.join(current_app.root_path, 'datasets', 'templates', '감사체크리스트.xlsx')
    wb = openpyxl.load_workbook(tpl_path)
    ws = wb.active

    def safe_write(row, col, value):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(rng.min_row, rng.min_col).value = value
                return
        ws.cell(row=row, column=col).value = value

    # 1) 상단 정보 채우기
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for rng in ws.merged_cells.ranges:
        val = str(ws.cell(rng.min_row, rng.min_col).value or '')
        if '피감사팀' in val:
            safe_write(rng.min_row, rng.min_col, f"{company_name} / {site_name}")
        if '감사원' in val:
            safe_write(rng.min_row, rng.min_col, '시스템 자동감사')
        if '감사일자' in val:
            safe_write(rng.min_row, rng.min_col, now_str)

    # 2) 데이터 조회 및 판정 로직 실행
    images = get_images_for_site(company_name, site_name)
    site_info = get_site_info(company_name, site_name)
    iso_checks, iso_reasons = analyze_environmental_aspects(images, site_info)
    print(f"[DEBUG] iso_checks={iso_checks}")
    print(f"[DEBUG] iso_reasons={iso_reasons}")

    # 3) 헤더행 찾기
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if str(cell.value or '').strip() == '감사항목':
                header_row = cell.row
                break
        if header_row:
            break
    print(f"[DEBUG] header_row={header_row}")
    if header_row is None:
        raise RuntimeError('감사항목 헤더를 찾을 수 없습니다.')

    # 4) 컬럼 인덱스 찾기
    cols = {'question':None, 'pass':None, 'fail':None}
    for idx, cell in enumerate(ws[header_row], start=1):
        text = str(cell.value or '').replace(' ', '').strip()
        if text == '감사항목': cols['question'] = idx
        elif text == '적합': cols['pass'] = idx
        elif text == '부적합': cols['fail'] = idx
    print(f"[DEBUG] cols={cols}")
    if None in cols.values():
        raise RuntimeError(f"컬럼 인덱스 누락: {cols}")

    # 5) 질문 목록
    iso_questions = [
        '환경측면을 파악하기 위한 분야별 주관 부서는 설정되었는가?',
        '환경측면/영향조사표 및 환경영향등록부가 기록되고 기능별로 정리되어 있는가?',
        '환경측면과 관련된 환경영향은 누락 없이 파악되고 있는가?',
        '환경측면의 중요성에 대한 평가는 정해진 기준을 준수하는가?',
        '중요한 환경측면 정보를 최신 자료로 유지하고 있는가?',
        '중요한 환경측면과 관련된 활동 및 운영이 식별되고 관련절차가 수립되고 기록되고 있는가?',
        '외주업체와 계약자에게 관련된 요구사항을 전달하는 의사소통은 수립되어 있는가?',
        '절차에 운영기준은 명시되고 있는가?',
        '주요 환경특성을 파악하고 모니터링 하고 있는가?',
        '사용하고 있는 모니터링 및 측정 장비를 교정, 검증하며 관련기록을 유지하는가?'
    ]

    # 6) 데이터 영역에 표시
    question_text_col = cols['question'] + 1  # 실제 질문 텍스트는 헤더 다음 열에 위치
    print(f"[DEBUG] question_text_col={question_text_col}")
    for r in range(header_row+1, ws.max_row+1):
        question_cell = ws.cell(row=r, column=question_text_col)
        q_text = str(question_cell.value or '').strip()
        print(f"[DEBUG] Row {r} question='{q_text}'")
        if q_text in iso_questions:
            idx = iso_questions.index(q_text)
            passed = iso_checks[idx]
            col_to_mark = cols['pass'] if passed else cols['fail']
            print(f"[DEBUG] marking idx={idx} passed={passed} at row={r}, col={col_to_mark}")
            safe_write(r, col_to_mark, '○')

    # 7) 수동확인필요 시트 작성
    manual_ws = wb.create_sheet('수동확인필요')
    manual_ws.append(['번호','감사항목','판정','판단사유'])
    for i, (ok, note) in enumerate(zip(iso_checks, iso_reasons), start=1):
        if not ok:
            manual_ws.append([i, iso_questions[i-1], '부적합', note])

    # 8) 출력
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def get_all_waste_objects():
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute("SELECT object_name FROM waste_objects")
        rows = cursor.fetchall()
    conn.close()
    return [row['object_name'] for row in rows]
